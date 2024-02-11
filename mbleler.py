import openpyxl as op
from openpyxl.styles import PatternFill as pf
import sys
from termcolor import colored as cl
import os

class mblelerClub:
    def __init__(self, kol, tmp, biasa, apik, elek):
        self.biasa = biasa
        self.apik = apik
        self.elek = elek
        self.kol = kol
        self.tmp = tmp

    def proses(self):
        try:
            def logo():
                print(cl("\n   ______ ___   __    _____ __    _____ _____\n  /  /  // _ \\ / /   / ___// /   / ___// __  \\\n / / / // _--// /__ /  __// /__ / ___// __  _/\n/_/_/_//____//____//____//____//____//_/  /_/","magenta"))
                print(cl("       --- (Auto Labeling by CukiD) --- \n","green"))

            if len(sys.argv) < 2:
                logo()
                print(cl("Untuk bantuan:\n  python.exe mbleler.py -h/--help\n","yellow"))
            elif len(sys.argv) == 2 and str(sys.argv[1]) == "-h" or str(sys.argv[1]) == "--help":
                logo()
                print("""Info:
    file   - file target (.xlsx)
    kolom  - head table target yang akan di lakukan eksekusi
    target - kata/kalimat yang mau diberi label
    label  - untuk memberi patternfill pada row yang mengandung kalimat tersebut
    output - file output (.xlsx)

Jenis label dan warnanya:
    negatif - merah
    positif - hijau
    netral  - abu-abu

Penggunaan:
    python3 mbleler.py [file] [kolom] [target] [label] [output]

Contoh:
    Pencarian kata:
        python3 mbleler.py data.xlsx biasa bangsat negatif outputnya.xlsx

    Pencarian kalimat:
        python3 mbleler.py data.xlsx opini '2024 siapapun hokagenya, konoha tetap berjaya' positif outputnya.xlsx\n""")
                
            elif len(sys.argv) == 6 and '.xlsx' in str(sys.argv[1]).lower() and ('positif' in str(sys.argv[-2]).lower() or 'negatif' in str(sys.argv[-2]).lower() or 'netral' in str(sys.argv[-2]).lower()) and '.xlsx' in str(sys.argv[-1]).lower():
                filenya = sys.argv[1]
                head = sys.argv[2]
                target = sys.argv[3]
                label = sys.argv[-2]
                output = sys.argv[-1]

                for aa in sys.argv:
                    self.tmp.append(aa)
                for _ in range(3):
                    self.tmp.pop(0)
                for _ in range(2):
                    self.tmp.pop()
                
                ini = ' '.join(self.tmp)

                if len(ini.split(' ')) == 1:
                    print(cl(f"\nProses mencari text {target}","green"))
                else:
                    print(cl(f"\nProses mencari kalimat {ini}","green"))

                wb = op.load_workbook(filenya)
                ws = wb.active
                max = ws.max_row - 1
                a = 0
                for aa in ws:
                    a += 1
                    if a == 1:
                        for bb in aa:
                            u = bb.value
                            p = bb.coordinate
                            if u == head:
                                o = 1
                                for _ in range(max):
                                    o += 1
                                    self.kol.append(f"{p[:1]}{o}")

                uu = 0
                for k in self.kol:
                    x = str(ws[k].value)
                    if target in x:
                        uu += 1
                        if label.lower() == 'positif':
                            n = pf(patternType='solid', fgColor=self.apik)
                            ws[k].fill = n
                        elif label.lower() == 'negatif':
                            n = pf(patternType='solid', fgColor=self.elek)
                            ws[k].fill = n
                        elif label.lower() == 'netral':
                            n = pf(patternType='solid', fgColor=self.biasa)
                            ws[k].fill = n
                        else:
                            print("Status tidak sesuai")

                print(f"\nTotal {head} ditemukan {uu} baris\n")

                no = 1
                if os.path.exists(output):
                    nf = output.split('.')
                    while True:
                        fileout = f"{nf[0]}{no}.{nf[1]}"
                        if not os.path.exists(fileout):
                            wb.save(f"{nf[0]}{no}.{nf[1]}")
                            print(cl(f"Silahkan cek file output {fileout}\n","green"))
                            break
                        no += 1
                else:
                    wb.save(output)
                    print(cl(f"Silahkan cek file output {output}\n","green"))
            else:
                print(cl("Untuk bantuan -h/--help\n","yellow"))
        except KeyboardInterrupt:
            print(cl("Proses dicancel ya\n","yellow"))

if __name__ == "__main__":
    netral = 'BCBCBC'
    positif = '00FF00'
    negatif = 'FF0000'
    kolom = []
    temp = []
    mbleler = mblelerClub(kolom, temp, netral, positif, negatif)
    mbleler.proses()
